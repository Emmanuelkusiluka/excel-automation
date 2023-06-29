import openpyxl as xl

from openpyxl.chart import Reference, PieChart


#kitchen


#foya kitchen


def foyakitchen(filename='Drop/kitchens/foyakitchen.xlsx'):
    sum_total = 0
    
    wb = xl.load_workbook(filename)
    sheet = wb['KITCHEN STOCK SHEETS (2)']
    
    # bites
    for row in range(6, 12+1):
        # items = sheet.cell(row, 2)
        # print(items.value)
        openingstock = sheet.cell(row, 4)
        addedstock =sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        #sold column
        sold = sheet.cell(row, 9)
        sellingprice= sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        #total sales calculations 
        totalsales = sellingprice.value * sold.value
        totalsales_cell.value = totalsales
                
        sum_total = sum_total + totalsales_cell.value

        
        # for num in int(totalsales_cell):
            # print(num)
        
        c_stock = sheet.cell(row, 11)    
        closing_stock = total - sold.value
        c_stock.value = closing_stock
    
    #tea
    for row in range(14, 17+1):
        # items = sheet.cell(row, 2)
        # print(items.value)
        openingstock = sheet.cell(row, 4)
        addedstock =sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        #sold column
        sold = sheet.cell(row, 9)
        sellingprice= sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        #total sales calculations 
        totalsales = sellingprice.value * sold.value
        totalsales_cell.value = totalsales
        
        sum_total = sum_total + totalsales_cell.value
        
        c_stock = sheet.cell(row, 11)    
        closing_stock = total - sold.value
        c_stock.value = closing_stock    
                      
    #egg   
    for row in range(19, 24+1):
        # items = sheet.cell(row, 2)
        # print(items.value)
        openingstock = sheet.cell(row, 4)
        addedstock =sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        #sold column
        sold = sheet.cell(row, 9)
        sellingprice= sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        #total sales calculations 
        totalsales = sellingprice.value * sold.value
        totalsales_cell.value = totalsales
        
        sum_total = sum_total + totalsales_cell.value
        
        c_stock = sheet.cell(row, 11)    
        closing_stock = total - sold.value
        c_stock.value = closing_stock
                 
        
    # chicken
    for row in range(19, 24+1):
        # items = sheet.cell(row, 2)
        # print(items.value)
        openingstock = sheet.cell(row, 4)
        addedstock =sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        #sold column
        sold = sheet.cell(row, 9)
        sellingprice= sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        #total sales calculations 
        totalsales = sellingprice.value * sold.value
        totalsales_cell.value = totalsales
        
        sum_total = sum_total + totalsales_cell.value
        
        c_stock = sheet.cell(row, 11)    
        closing_stock = total - sold.value
        c_stock.value = closing_stock
        
             
    #local chicken    
    for row in range(26, 26+1):
        # items = sheet.cell(row, 2)
        # print(items.value)
        openingstock = sheet.cell(row, 4)
        addedstock =sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        #sold column
        sold = sheet.cell(row, 9)
        sellingprice= sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        #total sales calculations 
        totalsales = sellingprice.value * sold.value
        totalsales_cell.value = totalsales
        
        sum_total = sum_total + totalsales_cell.value
        
        c_stock = sheet.cell(row, 11)    
        closing_stock = total - sold.value
        c_stock.value = closing_stock
        
           
    #broiler chicken   
    for row in range(28, 28+1):
        # items = sheet.cell(row, 2)
        # print(items.value)
        openingstock = sheet.cell(row, 4)
        addedstock =sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        #sold column
        sold = sheet.cell(row, 9)
        sellingprice= sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        #total sales calculations 
        totalsales = sellingprice.value * sold.value
        totalsales_cell.value = totalsales
        
        sum_total = sum_total + totalsales_cell.value
        
        c_stock = sheet.cell(row, 11)    
        closing_stock = total - sold.value
        c_stock.value = closing_stock  
        

    #rice and ugali 
    for row in range(30, 39+1):
        # items = sheet.cell(row, 2)
        # print(items.value)
        openingstock = sheet.cell(row, 4)
        addedstock =sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        #sold column
        sold = sheet.cell(row, 9)
        sellingprice= sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        #total sales calculations 
        totalsales = sellingprice.value * sold.value
        totalsales_cell.value = totalsales
        
        sum_total = sum_total + totalsales_cell.value
        
        c_stock = sheet.cell(row, 11)    
        closing_stock = total - sold.value
        c_stock.value = closing_stock
         
        
    # rice and pilau
    for row in range(41, 49+1):
        # items = sheet.cell(row, 2)
        # print(items.value)
        openingstock = sheet.cell(row, 4)
        addedstock =sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        #sold column
        sold = sheet.cell(row, 9)
        sellingprice= sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        #total sales calculations 
        totalsales = sellingprice.value * sold.value
        totalsales_cell.value = totalsales
        
        sum_total = sum_total + totalsales_cell.value
        
        c_stock = sheet.cell(row, 11)    
        closing_stock = total - sold.value
        c_stock.value = closing_stock       
        
    
    # Other sea food
    for row in range(51, 58):
        # items = sheet.cell(row, 2)
        # print(items.value)
        openingstock = sheet.cell(row, 4)
        addedstock =sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        #sold column
        sold = sheet.cell(row, 9)
        sellingprice= sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        #total sales calculations 
        totalsales = sellingprice.value * sold.value
        totalsales_cell.value = totalsales
        
        sum_total = sum_total + totalsales_cell.value
    
        
        c_stock = sheet.cell(row, 11)    
        closing_stock = total - sold.value
        c_stock.value = closing_stock



    for row in range(58, 58+1 ):
        # items = sheet.cell(row, 2)
        # print(items.value)
        openingstock = sheet.cell(row, 4)
        addedstock =sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        #sold column
        sold = sheet.cell(row, 9)
        sellingprice= sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        #total sales calculations 
        totalsales = sellingprice.value * sold.value
        totalsales_cell.value = totalsales
        
        sum_total = sum_total + totalsales_cell.value
        
        final_total_Cell= sheet
        final_total = sum_total
        
        final_total_Cell = final_total
        
        print(f"Foyakitchen Total sales: {final_total_Cell}")
        # print(final_total)
        
        c_stock = sheet.cell(row, 11)    
        closing_stock = total - sold.value
        c_stock.value = closing_stock       
        
             
    # values = Reference(sheet,
    #                    min_row=6,
    #                    max_row=58,
    #                    min_col=10,
    #                    max_col=10)

    # chart = PieChart()
    # chart.add_data(values)
    # sheet.add_chart(chart, 'c62')
    
   

    wb.save('Processed/kitchens/foyakitchen.xlsx')


#pool kitchen 
def poolkitchen(filename='Drop/kitchens/poolkitchen.xlsx'):

    sum_total = 0

    wb = xl.load_workbook(filename)
    sheet = wb['KITCHEN STOCK SHEETS (2)']

    # bites
    for row in range(6, 12 + 1):
        # items = sheet.cell(row, 2)
        # print(items.value)
        openingstock = sheet.cell(row, 4)
        addedstock = sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        # sold column
        sold = sheet.cell(row, 9)
        sellingprice = sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        # total sales calculations
        totalsales = sellingprice.value * sold.value
        totalsales_cell.value = totalsales

        sum_total = sum_total + totalsales_cell.value

        # for num in int(totalsales_cell):
        # print(num)

        c_stock = sheet.cell(row, 11)
        closing_stock = total - sold.value
        c_stock.value = closing_stock

    # tea
    for row in range(14, 17 + 1):
        # items = sheet.cell(row, 2)
        # print(items.value)
        openingstock = sheet.cell(row, 4)
        addedstock = sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        # sold column
        sold = sheet.cell(row, 9)
        sellingprice = sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        # total sales calculations
        totalsales = sellingprice.value * sold.value
        totalsales_cell.value = totalsales

        sum_total = sum_total + totalsales_cell.value

        c_stock = sheet.cell(row, 11)
        closing_stock = total - sold.value
        c_stock.value = closing_stock

        # egg
    for row in range(19, 24 + 1):
        # items = sheet.cell(row, 2)
        # print(items.value)
        openingstock = sheet.cell(row, 4)
        addedstock = sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        # sold column
        sold = sheet.cell(row, 9)
        sellingprice = sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        # total sales calculations
        totalsales = sellingprice.value * sold.value
        totalsales_cell.value = totalsales

        sum_total = sum_total + totalsales_cell.value

        c_stock = sheet.cell(row, 11)
        closing_stock = total - sold.value
        c_stock.value = closing_stock

    # chicken
    for row in range(19, 24 + 1):
        # items = sheet.cell(row, 2)
        # print(items.value)
        openingstock = sheet.cell(row, 4)
        addedstock = sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        # sold column
        sold = sheet.cell(row, 9)
        sellingprice = sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        # total sales calculations
        totalsales = sellingprice.value * sold.value
        totalsales_cell.value = totalsales

        sum_total = sum_total + totalsales_cell.value

        c_stock = sheet.cell(row, 11)
        closing_stock = total - sold.value
        c_stock.value = closing_stock

    # local chicken
    for row in range(26, 26 + 1):
        # items = sheet.cell(row, 2)
        # print(items.value)
        openingstock = sheet.cell(row, 4)
        addedstock = sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        # sold column
        sold = sheet.cell(row, 9)
        sellingprice = sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        # total sales calculations
        totalsales = sellingprice.value * sold.value
        totalsales_cell.value = totalsales

        sum_total = sum_total + totalsales_cell.value

        c_stock = sheet.cell(row, 11)
        closing_stock = total - sold.value
        c_stock.value = closing_stock

    # broiler chicken
    for row in range(28, 28 + 1):
        # items = sheet.cell(row, 2)
        # print(items.value)
        openingstock = sheet.cell(row, 4)
        addedstock = sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        # sold column
        sold = sheet.cell(row, 9)
        sellingprice = sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        # total sales calculations
        totalsales = sellingprice.value * sold.value
        totalsales_cell.value = totalsales

        sum_total = sum_total + totalsales_cell.value

        c_stock = sheet.cell(row, 11)
        closing_stock = total - sold.value
        c_stock.value = closing_stock

        # rice and ugali
    for row in range(30, 39 + 1):
        # items = sheet.cell(row, 2)
        # print(items.value)
        openingstock = sheet.cell(row, 4)
        addedstock = sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        # sold column
        sold = sheet.cell(row, 9)
        sellingprice = sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        # total sales calculations
        totalsales = sellingprice.value * sold.value
        totalsales_cell.value = totalsales

        sum_total = sum_total + totalsales_cell.value

        c_stock = sheet.cell(row, 11)
        closing_stock = total - sold.value
        c_stock.value = closing_stock

    # rice and pilau
    for row in range(41, 49 + 1):
        # items = sheet.cell(row, 2)
        # print(items.value)
        openingstock = sheet.cell(row, 4)
        addedstock = sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        # sold column
        sold = sheet.cell(row, 9)
        sellingprice = sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        # total sales calculations
        totalsales = sellingprice.value * sold.value
        totalsales_cell.value = totalsales

        sum_total = sum_total + totalsales_cell.value

        c_stock = sheet.cell(row, 11)
        closing_stock = total - sold.value
        c_stock.value = closing_stock

        # Other sea food
    for row in range(51, 58):
        # items = sheet.cell(row, 2)
        # print(items.value)
        openingstock = sheet.cell(row, 4)
        addedstock = sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        # sold column
        sold = sheet.cell(row, 9)
        sellingprice = sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        # total sales calculations
        totalsales = sellingprice.value * sold.value
        totalsales_cell.value = totalsales

        sum_total = sum_total + totalsales_cell.value

        c_stock = sheet.cell(row, 11)
        closing_stock = total - sold.value
        c_stock.value = closing_stock

    for row in range(58, 58 + 1):
        # items = sheet.cell(row, 2)
        # print(items.value)
        openingstock = sheet.cell(row, 4)
        addedstock = sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        # sold column
        sold = sheet.cell(row, 9)
        sellingprice = sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        # total sales calculations
        totalsales = sellingprice.value * sold.value
        totalsales_cell.value = totalsales

        sum_total = sum_total + totalsales_cell.value
        final_total_Cell = sheet
        final_total = sum_total

        final_total_Cell = final_total

        print(f"poolkitchen Total sales: {final_total_Cell}")
        # print(final_total)

        c_stock = sheet.cell(row, 11)
        closing_stock = total - sold.value
        c_stock.value = closing_stock

        # values = Reference(sheet,
    #                    min_row=6,
    #                    max_row=58,
    #                    min_col=10,
    #                    max_col=10)

    # chart = PieChart()
    # chart.add_data(values)
    # sheet.add_chart(chart, 'c62')

    wb.save('Processed/kitchens/poolkitchen.xlsx')
    
    
    
# Bar spreadsheets

#foyaBar
def foyaBar(filename='Drop/Bars/foyabar.xlsx'):
    wb = xl.load_workbook(filename)
    sheet = wb['sheet 1']
    
    sum_total = 0
    
    # my loops starts here 
    
    
    for row in range(7, 16+1):
        openingstock = sheet.cell(row, 4)
        addedstock =sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        #sold column
        sold = sheet.cell(row, 9)
        sellingprice= sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        #total sales calculations 
        totalsales = sellingprice.value * sold.value    
        totalsales_cell.value = totalsales
        
        c_stock = sheet.cell(row, 11)    
        closing_stock = total - sold.value
        c_stock.value = closing_stock
        
        sum_total = sum_total + totalsales_cell.value
        
    
        
        
    for row in range(18, 21+1):
        openingstock = sheet.cell(row, 4)
        addedstock =sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        #sold column
        sold = sheet.cell(row, 9)
        sellingprice= sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        #total sales calculations 
        totalsales = sellingprice.value * sold.value    
        totalsales_cell.value = totalsales
        
        c_stock = sheet.cell(row, 11)    
        closing_stock = total - sold.value
        c_stock.value = closing_stock
        
        sum_total = sum_total + totalsales_cell.value
    
    
    
    
    for row in range(23, 26+1):
        openingstock = sheet.cell(row, 4)
        # print(openingstock.value)
        addedstock =sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        #sold column
        sold = sheet.cell(row, 9)
        sellingprice= sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        #total sales calculations 
        totalsales = sellingprice.value * sold.value    
        totalsales_cell.value = totalsales
        
        c_stock = sheet.cell(row, 11)    
        closing_stock = total - sold.value
        c_stock.value = closing_stock
        
        sum_total = sum_total + totalsales_cell.value
    
        
    
    for row in range(28, 31+1):
        openingstock = sheet.cell(row, 4)
        addedstock =sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        #sold column
        sold = sheet.cell(row, 9)
        sellingprice= sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        #total sales calculations 
        totalsales = sellingprice.value * sold.value    
        totalsales_cell.value = totalsales
        
        c_stock = sheet.cell(row, 11)    
        closing_stock = total - sold.value
        c_stock.value = closing_stock
        
        sum_total = sum_total + totalsales_cell.value
    
        
        
        
    for row in range(33,46+1):
        openingstock = sheet.cell(row, 4)
        addedstock =sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        #sold column
        sold = sheet.cell(row, 9)
        sellingprice= sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        #total sales calculations 
        totalsales = sellingprice.value * sold.value    
        totalsales_cell.value = totalsales
        
        c_stock = sheet.cell(row, 11)    
        closing_stock = total - sold.value
        c_stock.value = closing_stock
        
        sum_total = sum_total + totalsales_cell.value
    
        
        
    for row in range(48, 52+1):
        openingstock = sheet.cell(row, 4)
        addedstock =sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        #sold column
        sold = sheet.cell(row, 9)
        sellingprice= sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        #total sales calculations 
        totalsales = sellingprice.value * sold.value    
        totalsales_cell.value = totalsales
        
        c_stock = sheet.cell(row, 11)    
        closing_stock = total - sold.value
        c_stock.value = closing_stock
        
        sum_total = sum_total + totalsales_cell.value
    
        
        
        
                 
    for row in range(54, 59+1):
        openingstock = sheet.cell(row, 4)
        addedstock =sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        #sold column
        sold = sheet.cell(row, 9)
        sellingprice= sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        #total sales calculations 
        totalsales = sellingprice.value * sold.value    
        totalsales_cell.value = totalsales
        
        c_stock = sheet.cell(row, 11)    
        closing_stock = total - sold.value
        c_stock.value = closing_stock
    
        sum_total = sum_total + totalsales_cell.value
        
        
    
    
    for row in range (61, 66+1):
        openingstock = sheet.cell(row, 4)
        addedstock =sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        #sold column
        sold = sheet.cell(row, 9)
        sellingprice= sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        #total sales calculations 
        totalsales = sellingprice.value * sold.value    
        totalsales_cell.value = totalsales
        
        c_stock = sheet.cell(row, 11)    
        closing_stock = total - sold.value
        c_stock.value = closing_stock
        
        sum_total = sum_total + totalsales_cell.value
      
        
    for row in range (68, 100+1):    
        openingstock = sheet.cell(row, 4)
        addedstock =sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        #sold column
        sold = sheet.cell(row, 9)
        sellingprice= sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        #total sales calculations 
        totalsales = sellingprice.value * sold.value    
        totalsales_cell.value = totalsales
        
        c_stock = sheet.cell(row, 11)    
        closing_stock = total - sold.value
        c_stock.value = closing_stock
        
        sum_total = sum_total + totalsales_cell.value
    
    
    for row in range(102, 116+1):
        openingstock = sheet.cell(row, 4)
        addedstock =sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        #sold column
        sold = sheet.cell(row, 9)
        sellingprice= sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        #total sales calculations 
        totalsales = sellingprice.value * sold.value    
        totalsales_cell.value = totalsales
        
        c_stock = sheet.cell(row, 11)    
        closing_stock = total - sold.value
        c_stock.value = closing_stock
        
        sum_total = sum_total + totalsales_cell.value
           
    
    for row in range(118, 119+1):    
        openingstock = sheet.cell(row, 4)
        addedstock =sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        #sold column
        sold = sheet.cell(row, 9)
        sellingprice= sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        #total sales calculations 
        totalsales = sellingprice.value * sold.value    
        totalsales_cell.value = totalsales
        
        c_stock = sheet.cell(row, 11)    
        closing_stock = total - sold.value
        c_stock.value = closing_stock
        
        sum_total = sum_total + totalsales_cell.value
    
    
    for row in range(121, 123+1):
        openingstock = sheet.cell(row, 4)
        addedstock =sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        #sold column
        sold = sheet.cell(row, 9)
        sellingprice= sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        #total sales calculations 
        totalsales = sellingprice.value * sold.value    
        totalsales_cell.value = totalsales
        
        c_stock = sheet.cell(row, 11)    
        closing_stock = total - sold.value
        c_stock.value = closing_stock
        
        sum_total = sum_total + totalsales_cell.value
              
        
    for row in range(125, 152+1):                    
        openingstock = sheet.cell(row, 4)
        addedstock =sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        #sold column
        sold = sheet.cell(row, 9)
        sellingprice= sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        #total sales calculations 
        totalsales = sellingprice.value * sold.value    
        totalsales_cell.value = totalsales
        
        c_stock = sheet.cell(row, 11)    
        closing_stock = total - sold.value
        c_stock.value = closing_stock
        
        sum_total = sum_total + totalsales_cell.value
        
          
    for row in range(153, 153+1):                    
        openingstock = sheet.cell(row, 4)
        addedstock =sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        #sold column
        sold = sheet.cell(row, 9)
        sellingprice= sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        #total sales calculations 
        totalsales = sellingprice.value * sold.value    
        totalsales_cell.value = totalsales
        
        c_stock = sheet.cell(row, 11)    
        closing_stock = total - sold.value
        c_stock.value = closing_stock
        
        sum_total = sum_total + totalsales_cell.value
        
        print(f"foyabar Total sales:{sum_total}")    
    
        
    values = Reference(sheet,
                       min_row=6,
                       max_row=58,
                       min_col=10,
                       max_col=10)

    chart = PieChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'G166')

    wb.save('Processed/bars/foyabar.xlsx')    
 
#poolbar   
def poolBar(filename='Drop/Bars/poolbar.xlsx'):
    wb = xl.load_workbook(filename)
    sheet = wb['sheet 1']
   
    sum_total = 0
    
    # my loops starts here 
    
    
    for row in range(7, 16+1):
        openingstock = sheet.cell(row, 4)
        addedstock =sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        #sold column
        sold = sheet.cell(row, 9)
        sellingprice= sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        #total sales calculations 
        totalsales = sellingprice.value * sold.value    
        totalsales_cell.value = totalsales
        
        c_stock = sheet.cell(row, 11)    
        closing_stock = total - sold.value
        c_stock.value = closing_stock
        
        sum_total = sum_total + totalsales_cell.value
        
    
        
        
    for row in range(18, 21+1):
        openingstock = sheet.cell(row, 4)
        addedstock =sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        #sold column
        sold = sheet.cell(row, 9)
        sellingprice= sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        #total sales calculations 
        totalsales = sellingprice.value * sold.value    
        totalsales_cell.value = totalsales
        
        c_stock = sheet.cell(row, 11)    
        closing_stock = total - sold.value
        c_stock.value = closing_stock
        
        sum_total = sum_total + totalsales_cell.value
    
    
    for row in range(23, 26+1):
        openingstock = sheet.cell(row, 4)
        # print(openingstock.value)
        addedstock =sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        #sold column
        sold = sheet.cell(row, 9)
        sellingprice= sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        #total sales calculations 
        totalsales = sellingprice.value * sold.value    
        totalsales_cell.value = totalsales
        
        c_stock = sheet.cell(row, 11)    
        closing_stock = total - sold.value
        c_stock.value = closing_stock
        
        sum_total = sum_total + totalsales_cell.value
    
         
    for row in range(28, 31+1):
        openingstock = sheet.cell(row, 4)
        addedstock =sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        #sold column
        sold = sheet.cell(row, 9)
        sellingprice= sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        #total sales calculations 
        totalsales = sellingprice.value * sold.value    
        totalsales_cell.value = totalsales
        
        c_stock = sheet.cell(row, 11)    
        closing_stock = total - sold.value
        c_stock.value = closing_stock
        
        sum_total = sum_total + totalsales_cell.value
    
              
    for row in range(33,46+1):
        openingstock = sheet.cell(row, 4)
        addedstock =sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        #sold column
        sold = sheet.cell(row, 9)
        sellingprice= sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        #total sales calculations 
        totalsales = sellingprice.value * sold.value    
        totalsales_cell.value = totalsales
        
        
        c_stock = sheet.cell(row, 11)    
        closing_stock = total - sold.value
        c_stock.value = closing_stock
        
        sum_total = sum_total + totalsales_cell.value
    
              
    for row in range(48, 52+1):
        openingstock = sheet.cell(row, 4)
        addedstock =sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        #sold column
        sold = sheet.cell(row, 9)
        sellingprice= sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        #total sales calculations 
        totalsales = sellingprice.value * sold.value    
        totalsales_cell.value = totalsales
        
        c_stock = sheet.cell(row, 11)    
        closing_stock = total - sold.value
        c_stock.value = closing_stock
        
        sum_total = sum_total + totalsales_cell.value
    
                      
    for row in range(54, 59+1):
        openingstock = sheet.cell(row, 4)
        addedstock =sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        #sold column
        sold = sheet.cell(row, 9)
        sellingprice= sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        #total sales calculations 
        totalsales = sellingprice.value * sold.value    
        totalsales_cell.value = totalsales
        
        c_stock = sheet.cell(row, 11)    
        closing_stock = total - sold.value
        c_stock.value = closing_stock
    
        sum_total = sum_total + totalsales_cell.value
        
          
    for row in range (61, 66+1):
        openingstock = sheet.cell(row, 4)
        addedstock =sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        #sold column
        sold = sheet.cell(row, 9)
        sellingprice= sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        #total sales calculations 
        totalsales = sellingprice.value * sold.value    
        totalsales_cell.value = totalsales
        
        c_stock = sheet.cell(row, 11)    
        closing_stock = total - sold.value
        c_stock.value = closing_stock
        
        sum_total = sum_total + totalsales_cell.value

            
    for row in range (68, 100+1):    
        openingstock = sheet.cell(row, 4)
        addedstock =sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        #sold column
        sold = sheet.cell(row, 9)
        sellingprice= sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        #total sales calculations 
        totalsales = sellingprice.value * sold.value    
        totalsales_cell.value = totalsales
        
        c_stock = sheet.cell(row, 11)    
        closing_stock = total - sold.value
        c_stock.value = closing_stock
        
        sum_total = sum_total + totalsales_cell.value
    
    
    for row in range(102, 116+1):
        openingstock = sheet.cell(row, 4)
        addedstock =sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        #sold column
        sold = sheet.cell(row, 9)
        sellingprice= sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        #total sales calculations 
        totalsales = sellingprice.value * sold.value    
        totalsales_cell.value = totalsales
        
        c_stock = sheet.cell(row, 11)    
        closing_stock = total - sold.value
        c_stock.value = closing_stock
        
        sum_total = sum_total + totalsales_cell.value
        
    
    for row in range(118, 119+1):    
        openingstock = sheet.cell(row, 4)
        addedstock =sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        #sold column
        sold = sheet.cell(row, 9)
        sellingprice= sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        #total sales calculations 
        totalsales = sellingprice.value * sold.value    
        totalsales_cell.value = totalsales
        
        c_stock = sheet.cell(row, 11)    
        closing_stock = total - sold.value
        c_stock.value = closing_stock
        
        sum_total = sum_total + totalsales_cell.value
    
    
    for row in range(121, 123+1):
        openingstock = sheet.cell(row, 4)
        addedstock =sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        #sold column
        sold = sheet.cell(row, 9)
        sellingprice= sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        #total sales calculations 
        totalsales = sellingprice.value * sold.value    
        totalsales_cell.value = totalsales
        
        c_stock = sheet.cell(row, 11)    
        closing_stock = total - sold.value
        c_stock.value = closing_stock
        
        sum_total = sum_total + totalsales_cell.value
        
            
    for row in range(125, 152+1):                    
        openingstock = sheet.cell(row, 4)
        addedstock =sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        #sold column
        sold = sheet.cell(row, 9)
        sellingprice= sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        #total sales calculations 
        totalsales = sellingprice.value * sold.value    
        totalsales_cell.value = totalsales
        
        c_stock = sheet.cell(row, 11)    
        closing_stock = total - sold.value
        c_stock.value = closing_stock
        
        sum_total = sum_total + totalsales_cell.value
        
      
    for row in range(153, 153+1):                    
        openingstock = sheet.cell(row, 4)
        addedstock =sheet.cell(row, 5)
        spoiled = sheet.cell(row, 6)
        # total column operations
        total = ((openingstock.value + addedstock.value) - spoiled.value)
        corrected_price_cell = sheet.cell(row, 7)
        corrected_price_cell.value = total
        #sold column
        sold = sheet.cell(row, 9)
        sellingprice= sheet.cell(row, 8)
        totalsales_cell = sheet.cell(row, 10)
        #total sales calculations 
        totalsales = sellingprice.value * sold.value    
        totalsales_cell.value = totalsales
        
        c_stock = sheet.cell(row, 11)    
        closing_stock = total - sold.value
        c_stock.value = closing_stock
        
        sum_total = sum_total + totalsales_cell.value
        
        print(f"pool bar Total sales: {sum_total} ")    
        
        
    # values = Reference(sheet,
    #                    min_row=6,
    #                    max_row=58,
    #                    min_col=10,
    #                    max_col=10)

    # chart = PieChart()
    # chart.add_data(values)
    # sheet.add_chart(chart, 'c62')

    wb.save('Processed/bars/poolbar.xlsx')    
            
    
    